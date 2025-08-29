
from openpyxl import load_workbook
from openpyxl.comments import Comment

HEADER_COMMENTS = {
    "Timestamp": "Zeitpunkt der Auswertung.",
    "Folder": "Ordner/Datensatz-ID.",
    "Version": "Pipeline-Version/Run.",
    "Typ": "Quelle: Python (eigene M3C2) oder CC (CloudCompare).",
    "Gesamt": "Anzahl aller Distanzwerte inkl. NaN. Nur der Kontext (Größe) – kein Qualitätsmaß.",
    "NaN": "Anzahl ohne gültiges Ergebnis (z.B. keine Nachbarn). Weniger ist besser.",
    "% NaN": "NaN/ Gesamt. Hoch = viele Nichttreffer/Abdeckungsprobleme.",
    "% Valid": "1 − %NaN. Hoch = gute Abdeckung/Robustheit.",
    "Valid Count": "Anzahl gültiger (geclippter) Werte. Vergleichbar zu %Valid.",
    "Valid Sum": "Σ der Distanzen (geclippt). Nahe 0 = wenig Bias. Vorzeichen beachten.",
    "Valid Squared Sum": "Σ der Quadrate (geclippt). Hoch = viel „Energie“/Streuung/Änderung.",
    "Normal Scale": "Radius zur Normalenschätzung. Zu klein = rausch-/instabil; zu groß = überglättet.",
    "Search Scale": "Such-/Projektionsradius (oft ≈2× Normal). Zu klein = weniger Treffer; zu groß = Glättung.",
    "Min": "Kleinstes (ggf. negatives) Distanz-Extrem. Nur Hinweis auf Ausreißer.",
    "Max": "Größtes Distanz-Extrem. Hoch = pot. Ausreißer/echte große Änderung.",
    "Mean": "Mittelwert (Bias). Ideal nahe 0. Vorzeichen = systematische Verschiebung.",
    "Median": "Median (robuster Lagewert). Näher an 0 bei schiefen/ausreißerbehafteten Daten.",
    "RMS": "Root-Mean-Square. Für biasfreie Daten ≈ Std; sonst RMS² = Std² + Bias².",
    "Std": "Standardabweichung. Niedrig = wenig Streuung/rauscharm. Sensibel für Ausreißer.",
    "MAE": "Mean Absolute Error = ⟨|x|⟩. Robust(er) als Std; für N(0,σ): MAE≈0.798·σ.",
    "MAE Inlier": "MAE nach Entfernen von Ausreißern (|x|>3·RMSE). Vergleichbar zwischen Runs.",
    "NMAD": "1.4826·MAD. Robuste σ-Schätzung; kaum ausreißerempfindlich.",
    "NMAD Inlier": "NMAD nur auf Inliern. Typisch ≤ Std und ≈ Std bei normalverteilten Daten.",
    "Outlier Count": "Anzahl |x|>3·RMSE. Hoch = schwere Tails/echte starke Änderungen.",
    "Inlier Count": "Anzahl |x|≤3·RMSE. Ergänzt sich zu Valid Count.",
    "Mean Inlier": "Mittelwert ohne Ausreißer. Gute Bias-Schätzung.",
    "Std Inlier": "Std ohne Ausreißer. Besser vergleichbar zwischen Datensätzen.",
    "Mean Outlier": "Mittelwert der Ausreißer. Quantisiert Vorzeichen/Tendenz der starken Abweichungen.",
    "Std Outlier": "Std der Ausreißer. Hoch = sehr heterogene Ausreißer.",
    "Pos Outlier": "Ausreißer >0. Überschuss → asymmetrische positive Tail.",
    "Neg Outlier": "Ausreißer <0. Überschuss → asymmetrische negative Tail.",
    "Pos Inlier": "Inlier >0. Vorzeichenverhältnis zeigt leichte Asymmetrien.",
    "Neg Inlier": "Inlier <0. Siehe Pos Inlier.",
    "Q05": "5%-Quantil. Zusammen mit Q95 Breite der Verteilung (robust).",
    "Q25": "25%-Quantil (unteres Quartil).",
    "Q75": "75%-Quantil (oberes Quartil).",
    "Q95": "95%-Quantil. Kleinere |Q95| besser (enger, weniger Ausreißer).",
    "Gauss Mean": "μ des Normalverteilungs-Fits. Sollte nahe 0 liegen (Bias).",
    "Gauss Std": "σ des Fits. Vergleichbar mit Std/NMAD.",
    "Gauss Chi2": "Pearson-Chi² gegen Gauß (nur Bins mit genügender Erwartung). Niedriger = besserer Fit; zum Vergleich zwischen Runs nutzen (absoluter Wert hängt von Binning/N ab).",
    "Weibull a": "Weibull-Form (shape). Klein (<1)=starke Tails; groß(>3)=„spitzer“/symmetrischer. Achtung: mit loc kann auch linksseitig „passen“.",
    "Weibull b": "Weibull-Skala (scale). Größer = breiter.",
    "Weibull shift": "loc-Parameter (Verschiebung).",
    "Weibull mode": "Modus der gefitteten Weibull. Lage der häufigsten Werte.",
    "Weibull skewness": "Schiefe des Weibull-Fits. >0: Rechts-Tail; <0: Links-Tail.",
    "Weibull Chi2": "Pearson-Chi² gegen Weibull. Niedriger = besserer Fit. Nur relativ vergleichen.",
    "Skewness": "Schiefe der Distanzverteilung. ~0 gut; >0 Rechts-Tail; <0 Links-Tail. |skew|>1 = starke Asymmetrie.",
    "Kurtosis": "Exzess-Kurtosis (Normal=0). >0 schwere Tails; <0 leichte Tails.",
    "Anteil |Distanz| > 0.01": "Fraktion außerhalb ±1 cm. Niedrig = gut (genauer).",
    "Anteil [-2Std,2Std]": "Anteil in [−2·σ,+2·σ] um 0 (σ=Gauss Std). Ideal ~95% (bei biasfrei normal). Bias senkt den Anteil.",
    "Max |Distanz|": "Größter Absolutwert. Nur Extrem-Hinweis.",
    "Bias": "Gleichbedeutend mit Mean (systematischer Versatz). |Bias| klein = gut.",
    "Within-Tolerance": "Anteil |x| ≤ Toleranz (Default 1 cm). Hoch = gut; gleicht 1 − Anteil>|tol|.",
    "ICC": "Intraklassenkorrelation – hier Platzhalter (nicht echte ICC-Berechnung).",
    "CCC": "Concordance Corr. Coef. – Achtung: hier nur vereinfachter Proxy, nicht der echte CCC.",
    "Bland-Altman Lower": "Bias − 1.96·σ: untere Übereinstimmungsgrenze.",
    "Bland-Altman Upper": "Bias + 1.96·σ: obere Übereinstimmungsgrenze.",
    "Jaccard Index": "In aktueller Implementierung identisch zu Within-Tolerance (ein-Set-Fall). Echte Overlap-Kennzahl nur beim Vergleich zweier Läufe.",
    "Dice Coefficient": "Wie Jaccard hier identisch zu Within-Tolerance (ein-Set-Fall).",
    "Distances Path": "Pfad der Distanz-Quelle (Nachvollziehbarkeit).",
    "Params Path": "Pfad der Normal/Search-Scale-Datei.",
}

def add_header_comments(
    xlsx_path: str,
    sheet_name: str = "Results",
    header_row: int = 1,
    author: str = "StatisticsService",
    overwrite: bool = True,
    box_width: float = 300,
    box_height: float = 160
):
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' nicht gefunden in {xlsx_path}")
    ws = wb[sheet_name]

    # Lese die Überschriftenzeile
    headers = [cell.value for cell in ws[header_row]]

    for col_idx, header in enumerate(headers, start=1):
        if not header:
            continue
        if header not in HEADER_COMMENTS:
            continue
        cell = ws.cell(row=header_row, column=col_idx)
        if (cell.comment is not None) and not overwrite:
            continue
        c = Comment(HEADER_COMMENTS[header], author)
        c.visible = False
        c.width = box_width
        c.height = box_height
        cell.comment = c

    wb.save(xlsx_path)

if __name__ == "__main__":
    add_header_comments("m3c2_stats_all.xlsx", sheet_name="Results")