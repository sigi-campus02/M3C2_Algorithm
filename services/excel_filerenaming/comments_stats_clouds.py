from openpyxl import load_workbook
from openpyxl.comments import Comment
import re

# ---- feste (exakte) Header-Kommentare für Single-Cloud-Stats ----
CLOUD_HEADER_COMMENTS = {
    "Timestamp": "Zeitpunkt der Auswertung.",
    "File/Folder": "Quelle der Punktwolke (Dateipfad oder Ordner).",
    "Role": "Rolle in einem Datensatz-Ordner (mov/ref). Leer bei Einzeldatei.",

    # Identifikation/Fläche/Dichte
    "Num Points": "Anzahl der Punkte (nach optionalem Subsample nur für lokale Metriken).",
    "Area Source": "Flächenquelle: 'bbox' (Bounding-Box), 'convex_hull' (konvexe Hülle) oder 'given' (vom Nutzer).",
    "Area XY [m^2]": "2D-Fläche der Wolke in XY – Basis für globale Flächendichte.",
    "Density Global [pt/m^2]": "Globale Flächendichte: Punkte pro Quadratmeter in XY.",

    # Globale Z-Statistik (roh)
    "Z Min": "Kleinster Z-Wert (roh).",
    "Z Max": "Größter Z-Wert (roh).",
    "Z Mean": "Mittlerer Z-Wert (roh).",
    "Z Median": "Median der Z-Werte (roh).",
    "Z Std": "Standardabweichung der Z-Werte (roh).",
    "Z Q05": "5%-Quantil der Z-Werte (roh).",
    "Z Q25": "25%-Quantil der Z-Werte (roh).",
    "Z Q75": "75%-Quantil der Z-Werte (roh).",
    "Z Q95": "95%-Quantil der Z-Werte (roh).",

    # 3D-Lokaldichte & Rauigkeit
    "Local Density Mean [pt/m^3]": "Mittlere lokale 3D-Dichte: Nachbarn pro Kugelvolumen (Radius=‘Radius [m]’).",
    "Local Density Median [pt/m^3]": "Median der lokalen 3D-Dichte.",
    "Local Density Q05 [pt/m^3]": "5%-Quantil der lokalen 3D-Dichte.",
    "Local Density Q95 [pt/m^3]": "95%-Quantil der lokalen 3D-Dichte.",

    "Roughness Mean [m]": "Mittlere lokale Rauigkeit: Std der orthogonalen Residuen zur PCA-Best-Fit-Ebene.",
    "Roughness Median [m]": "Median der lokalen Rauigkeit.",
    "Roughness Q05 [m]": "5%-Quantil der lokalen Rauigkeit.",
    "Roughness Q95 [m]": "95%-Quantil der lokalen Rauigkeit.",

    # PCA-Formmaße
    "Linearity Mean": "Mittlere Linearität: (λ1−λ2)/λ1. Nahe 1 = linienartig.",
    "Linearity Median": "Median der Linearität.",
    "Planarity Mean": "Mittlere Planarität: (λ2−λ3)/λ1. Nahe 1 = flächig.",
    "Planarity Median": "Median der Planarität.",
    "Sphericity Mean": "Mittlere Sphärizität: λ3/λ1. Nahe 1 = kugelig/rauschig.",
    "Sphericity Median": "Median der Sphärizität.",

    # Normalenkonsistenz
    "Normal Std Angle [deg]": "Winkel-Std der lokalen Normalen (nach Ausrichtung). Klein = konsistente Normale.",

    # Meta
    "Radius [m]": "Nachbarschaftsradius (Kugel) für lokale Metriken.",
    "k-NN": "k in k-NN für mittlere Abstandsmetriken.",
    "Sampled Points": "Anzahl Punkte im Subsample S für lokale Auswertungen.",
}

# ---- optionale / zukünftige Spalten (werden kommentiert, falls vorhanden) ----
CLOUD_HEADER_COMMENTS_OPTIONAL = {
    # Falls du detrended-Z einführst:
    "Z(detrended) Min": "Min der Höhen relativ zur globalen Best-Fit-Ebene.",
    "Z(detrended) Max": "Max der Höhen relativ zur globalen Best-Fit-Ebene.",
    "Z(detrended) Mean": "Mean relativ zur globalen Best-Fit-Ebene.",
    "Z(detrended) Median": "Median relativ zur globalen Best-Fit-Ebene.",
    "Z(detrended) Std": "Std relativ zur globalen Best-Fit-Ebene.",
    "Z(detrended) Q05": "5%-Quantil (detrended).",
    "Z(detrended) Q25": "25%-Quantil (detrended).",
    "Z(detrended) Q75": "75%-Quantil (detrended).",
    "Z(detrended) Q95": "95%-Quantil (detrended).",

    # Falls du lokale 2D-Dichte ergänzt:
    "Local Density2D Mean [pt/m^2]": "Mittlere lokale 2D-Dichte: Nachbarn pro Kreisfläche (XY, Radius=‘Radius [m]’).",
    "Local Density2D Median [pt/m^2]": "Median der lokalen 2D-Dichte.",
    "Local Density2D Q05 [pt/m^2]": "5%-Quantil der lokalen 2D-Dichte.",
    "Local Density2D Q95 [pt/m^2]": "95%-Quantil der lokalen 2D-Dichte.",

    # Nachbarschaftszahlen & Coverage:
    "Neighborhood Coverage": "Anteil der Punkte mit genügend Nachbarn (z. B. ≥3).",
    "Neighbors Mean": "Mittlere Nachbarschaftsgröße (Anzahl Nachbarn im Radius).",
    "Neighbors Median": "Median der Nachbarschaftsgröße.",
    "Neighbors Q05": "5%-Quantil der Nachbarschaftsgröße.",
    "Neighbors Q95": "95%-Quantil der Nachbarschaftsgröße.",

    # Completeness:
    "Completeness Threshold [pt/m^2]": "Schwellenwert für 2D-Dichte zu Lochdetektion (z. B. 25% der globalen Dichte).",
    "Hole Ratio (<tau)": "Anteil lokaler Bereiche unter Dichteschwelle (Niedrig = gute Abdeckung).",

    # Normalenqualität (falls ergänzt):
    "Spherical Variance of Normals": "1−R (Resultantenlänge); 0=perfekt konsistent, 1=zufällig.",
}

# ---- Muster (regex) für dynamische Header ----
CLOUD_PATTERN_COMMENTS = [
    # "Mean NN Dist (1..k)"
    (re.compile(r"^Mean NN Dist \(1\.\.\d+\)$"),
     "Mittlerer Abstand zu den 1..k nächsten Nachbarn (Subsample S). Proxy für typische Punktabstände."),
    # "Mean Dist to k-NN"
    (re.compile(r"^Mean Dist to \d+-NN$"),
     "Mittlerer Abstand zum k-ten Nachbarn (charakteristische Nachbarschaftsskala)."),
]

def add_cloud_header_comments(xlsx_path: str,
                              sheet_name: str = "CloudStats",
                              header_row: int = 1,
                              author: str = "StatisticsService",
                              overwrite: bool = True,
                              box_width: float = 300,
                              box_height: float = 160) -> None:
    """
    Hängt erklärende Kommentare an die Header der Single-Cloud-Statistik an.
    Unterstützt exakte Header und dynamische (regex-basierte) Header.
    """
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' nicht gefunden in {xlsx_path}")
    ws = wb[sheet_name]

    headers = [cell.value for cell in ws[header_row]]
    for col_idx, header in enumerate(headers, start=1):
        if not header or not isinstance(header, str):
            continue

        comment_text = None

        # exakte Treffer prüfen
        if header in CLOUD_HEADER_COMMENTS:
            comment_text = CLOUD_HEADER_COMMENTS[header]
        elif header in CLOUD_HEADER_COMMENTS_OPTIONAL:
            comment_text = CLOUD_HEADER_COMMENTS_OPTIONAL[header]
        else:
            # Muster prüfen
            for pattern, text in CLOUD_PATTERN_COMMENTS:
                if pattern.match(header):
                    comment_text = text
                    break

        if not comment_text:
            continue

        cell = ws.cell(row=header_row, column=col_idx)
        if (cell.comment is not None) and not overwrite:
            continue

        c = Comment(comment_text, author)
        c.visible = False
        c.width = box_width
        c.height = box_height
        cell.comment = c

    wb.save(xlsx_path)

# Beispielaufruf:
add_cloud_header_comments("cloud_stats.xlsx", sheet_name="CloudStats")
