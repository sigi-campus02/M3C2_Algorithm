from __future__ import annotations
import os
from typing import Dict, List, Optional, Tuple
import numpy as np
import pandas as pd
from scipy.stats import norm, weibull_min
from datetime import datetime
from sklearn.decomposition import PCA
from sklearn.neighbors import NearestNeighbors
from datasource.datasource import DataSource


# Kanonische Spaltenreihenfolge für Statistik-Exports
CANONICAL_COLUMNS = [
    "Timestamp", "Folder", "Version", "Total Points",
    "Normal Scale", "Search Scale",
    "NaN", "% NaN", "% Valid",
    "Valid Count", "Valid Sum", "Valid Squared Sum",
    "Valid Count Inlier", "Valid Sum Inlier", "Valid Squared Sum Inlier",
    "Min", "Max", "Mean", "Median", "RMS", "Std Empirical", "MAE", "NMAD",
    "Min Inlier", "Max Inlier", "Mean Inlier", "Median Inlier", "RMS Inlier",
    "Std Inlier", "MAE Inlier", "NMAD Inlier",
    "Outlier Multiplicator", "Outlier Threshold", "Outlier Method",
    "Inlier Count", "Pos Inlier", "Neg Inlier",
    "Pos Outlier", "Neg Outlier", "Outlier Count",
    "Mean Outlier", "Std Outlier",
    "Q05", "Q25", "Q75", "Q95", "IQR",
    "Q05 Inlier", "Q25 Inlier", "Q75 Inlier", "Q95 Inlier", "IQR Inlier",
    "Gauss Mean", "Gauss Std",
    "Weibull a", "Weibull b", "Weibull shift", "Weibull mode", "Weibull skewness",
    "Skewness", "Kurtosis",
    "Distances Path", "Params Path"
]


class StatisticsService:
    # =============================
    # Public API
    # =============================

    @staticmethod
    def calc_stats(
        distances: np.ndarray,
        params_path: Optional[str] = None,
        bins: int = 256,
        range_override: Optional[Tuple[float, float]] = None,
        min_expected: Optional[float] = None,
        tolerance: float = 0.01,
        outlier_multiplicator: float = 3.0,
        outlier_method: str = "rmse"

    ) -> Dict:
        """Berechne diverse Metriken aus den gegebenen Distanzwerten.

        Parameters
        ----------
        distances:
            Array mit Distanzwerten, das NaN enthalten kann.
        params_path:
            Optionaler Pfad zu einer Datei mit ``NormalScale`` und ``SearchScale`` Parametern.
        bins:
            Anzahl der Bins für das Histogramm.
        range_override:
            Optionales Tupel zur expliziten Festlegung des Wertebereichs.
        min_expected:
            Mindestanzahl erwarteter Werte pro Bin für die Chi²-Berechnung.
        tolerance:
            Toleranzschwelle zur Bewertung der Distanzwerte.
        """

        total_count = len(distances)
        nan_count = int(np.isnan(distances).sum())
        valid = distances[~np.isnan(distances)]
        if valid.size == 0:
            raise ValueError("No valid distances")

        # Range wie in CC
        if range_override is None:
            data_min, data_max = float(np.min(valid)), float(np.max(valid))
        else:
            data_min, data_max = map(float, range_override)

        # Clip (wie CC) für Histogramm + Fits
        clipped = valid[(valid >= data_min) & (valid <= data_max)]
        if clipped.size == 0:
            raise ValueError("All values fall outside the selected range")

        # Basis-Statistiken für alle Werte (inkl. Outlier)
        stats_all = StatisticsService._basic_stats(clipped, tolerance)
        valid_sum = stats_all["Valid Sum"]
        valid_squared_sum = stats_all["Valid Squared Sum"]
        avg = stats_all["Mean"]
        med = stats_all["Median"]
        rms = stats_all["RMS"]
        std_empirical = stats_all["Std Empirical"]
        mae = stats_all["MAE"]
        nmad = stats_all["NMAD"]

        # Histogramm
        hist, bin_edges = np.histogram(clipped, bins=bins, range=(data_min, data_max))
        hist = hist.astype(float)

        # Fits (Gauss & Weibull)
        fit_results = StatisticsService._fit_distributions(
            clipped, hist, bin_edges, min_expected
        )
        mu = fit_results["mu"]
        std = fit_results["std"]
        pearson_gauss = fit_results["pearson_gauss"]
        a = fit_results["a"]
        b = fit_results["b"]
        loc = fit_results["loc"]
        pearson_weib = fit_results["pearson_weib"]
        skew_weibull = fit_results["skew_weibull"]
        mode_weibull = fit_results["mode_weibull"]

        # Optional: CC-/Params-Datei
        normal_scale, search_scale = StatisticsService._load_params(params_path)

        # Outliers
        outlier_mask, outlier_threshold = StatisticsService.get_outlier_mask(clipped, outlier_method, outlier_multiplicator)
        inliers = clipped[~outlier_mask]
        outliers = clipped[outlier_mask]
        outlier_info = StatisticsService._compute_outliers(inliers, outliers)
        outlier_count = outlier_info["outlier_count"]
        inlier_count = outlier_info["inlier_count"]
        mean_out = outlier_info["mean_out"]
        std_out = outlier_info["std_out"]
        pos_out = outlier_info["pos_out"]
        neg_out = outlier_info["neg_out"]
        pos_in = outlier_info["pos_in"]
        neg_in = outlier_info["neg_in"]

        stats_in = StatisticsService._basic_stats(inliers, tolerance)
        mean_in = stats_in["Mean"]
        std_in = stats_in["Std Empirical"]
        mae_in = stats_in["MAE"]
        nmad_in = stats_in["NMAD"]
        min_in = stats_in["Min"]
        max_in = stats_in["Max"]
        median_in = stats_in["Median"]
        rms_in = stats_in["RMS"]
        q05_in = stats_in["Q05"]
        q25_in = stats_in["Q25"]
        q75_in = stats_in["Q75"]
        q95_in = stats_in["Q95"]
        iqr_in = stats_in["IQR"]
        skew_in = stats_in["Skewness"]
        kurt_in = stats_in["Kurtosis"]
        share_abs_gt_in = stats_in["Anteil |Distanz| > 0.01"]
        share_2std_in = stats_in["Anteil [-2Std,2Std]"]
        max_abs_in = stats_in["Max |Distanz|"]
        bias_in = stats_in["Bias"]
        within_tolerance_in = stats_in["Within-Tolerance"]
        jaccard_in = stats_in["Jaccard Index"]
        dice_in = stats_in["Dice Coefficient"]
        valid_count_in = stats_in["Valid Count"]
        valid_sum_in = stats_in["Valid Sum"]
        valid_squared_sum_in = stats_in["Valid Squared Sum"]

        # Bias & Toleranz
        bias = stats_all["Bias"]
        within_tolerance = stats_all["Within-Tolerance"]

        # ICC/CCC/Bland-Altman
        icc = np.nan  # Placeholder
        mean_dist = float(np.mean(clipped))
        std_dist = float(np.std(clipped))
        ccc = (2 * mean_dist * std_dist) / (mean_dist**2 + std_dist**2) if mean_dist != 0 else np.nan

        bland_altman_lower = bias - 1.96 * std_dist
        bland_altman_upper = bias + 1.96 * std_dist

        # Overlap (Jaccard/Dice)
        jaccard_index = stats_all["Jaccard Index"]
        dice_coefficient = stats_all["Dice Coefficient"]

        return {
            # 1) Counts & Scales
            "Total Points": total_count,
            "NaN": nan_count,
            "% NaN": (nan_count / total_count) if total_count > 0 else np.nan,
            "% Valid": (1 - nan_count / total_count) if total_count > 0 else np.nan,
            "Valid Count": int(clipped.size),
            "Valid Sum": valid_sum,
            "Valid Squared Sum": valid_squared_sum,
            "Valid Count Inlier": int(valid_count_in),
            "Valid Sum Inlier": valid_sum_in,
            "Valid Squared Sum Inlier": valid_squared_sum_in,
            "Normal Scale": normal_scale,
            "Search Scale": search_scale,

            # 2) Lage & Streuung
            "Min": float(np.nanmin(distances)),
            "Max": float(np.nanmax(distances)),
            "Mean": avg,
            "Median": med,
            "RMS": rms,
            "Std Empirical": std_empirical,
            "MAE": mae,
            "NMAD": nmad,
            "Min Inlier": min_in,
            "Max Inlier": max_in,
            "Mean Inlier": mean_in,
            "Median Inlier": median_in,
            "RMS Inlier": rms_in,
            "Std Inlier": std_in,
            "MAE Inlier": mae_in,
            "NMAD Inlier": nmad_in,

            # 3) Outlier / Inlier
            "Outlier Count": outlier_count,
            "Inlier Count": inlier_count,
            "Mean Outlier": mean_out,
            "Std Outlier": std_out,
            "Pos Outlier": pos_out,
            "Neg Outlier": neg_out,
            "Pos Inlier": pos_in,
            "Neg Inlier": neg_in,
            "Outlier Multiplicator": outlier_multiplicator,
            "Outlier Threshold": outlier_threshold,
            "Outlier Method": outlier_method,

            # 4) Quantile
            "Q05": stats_all["Q05"],
            "Q25": stats_all["Q25"],
            "Q75": stats_all["Q75"],
            "Q95": stats_all["Q95"],
            "IQR": stats_all["IQR"],
            "Q05 Inlier": q05_in,
            "Q25 Inlier": q25_in,
            "Q75 Inlier": q75_in,
            "Q95 Inlier": q95_in,
            "IQR Inlier": iqr_in,

            # 5) Fit-Metriken
            "Gauss Mean": float(mu),
            "Gauss Std": float(std),
            "Gauss Chi2": float(pearson_gauss),
            "Weibull a": float(a),
            "Weibull b": float(b),
            "Weibull shift": float(loc),
            "Weibull mode": mode_weibull,
            "Weibull skewness": skew_weibull,
            "Weibull Chi2": float(pearson_weib),

            # 6) Weitere Kennzahlen
            "Skewness": stats_all["Skewness"],
            "Kurtosis": stats_all["Kurtosis"],
            # "Skewness Inlier": skew_in,
            # "Kurtosis Inlier": kurt_in,
            # "Anteil |Distanz| > 0.01": stats_all["Anteil |Distanz| > 0.01"], # 1. Anteil |Distanz| > 0.01 (1 cm-Grenze)
            # "Anteil [-2Std,2Std]": stats_all["Anteil [-2Std,2Std]"], # 2. Anteil innerhalb ±2·Std
            # "Anteil |Distanz| > 0.01 Inlier": share_abs_gt_in,
            # "Anteil [-2Std,2Std] Inlier": share_2std_in,
            # "Max |Distanz|": stats_all["Max |Distanz|"], # 3. Maximaler Absolutwert (Extremabweichung)
            # "Bias": bias,
            # "Within-Tolerance": within_tolerance, # 4. Within-Tolerance (default: ±1 cm)
            # "Max |Distanz| Inlier": max_abs_in,
            # "Bias Inlier": bias_in,
            # "Within-Tolerance Inlier": within_tolerance_in,
            # "ICC": icc,
            # "CCC": ccc,
            # "Bland-Altman Lower": bland_altman_lower,
            # "Bland-Altman Upper": bland_altman_upper,
            # "Jaccard Index": jaccard_index,
            # "Dice Coefficient": dice_coefficient,
            # "Jaccard Index Inlier": jaccard_in,
            # "Dice Coefficient Inlier": dice_in
        }

    @classmethod
    def compute_m3c2_statistics(
        cls,
        folder_ids: List[str],
        filename_ref: str = "",
        process_python_CC: str = "python",
        bins: int = 256,
        range_override: Optional[Tuple[float, float]] = None,
        min_expected: Optional[float] = None,
        out_path: str = "m3c2_stats_all.xlsx",
        sheet_name: str = "Results",
        output_format: str = "excel",
        outlier_multiplicator: float = 3.0,
        outlier_method: str = "rmse"
    ) -> pd.DataFrame:
        """
        Liest je Folder {version}_m3c2_distances.txt (Python) und optional CloudCompare
        und hängt die Ergebnisse an eine Datei (Excel oder JSON) an.
        Spalten: Folder | Version | Typ | ... (Metriken)
        """
        rows: List[Dict] = []

        for fid in folder_ids:
            # ----- Python
            if process_python_CC == "python":
                py_dist_path   = cls._resolve(fid, f"python_{filename_ref}_m3c2_distances.txt")
                py_params_path = cls._resolve(fid, f"python_{filename_ref}_m3c2_params.txt")
                if os.path.exists(py_dist_path):
                    values = np.loadtxt(py_dist_path)
                    stats = cls.calc_stats(
                        values,
                        params_path=py_params_path if os.path.exists(py_params_path) else None,
                        bins=bins,
                        range_override=range_override,
                        min_expected=min_expected,
                        outlier_multiplicator=outlier_multiplicator,
                        outlier_method=outlier_method
                    )
                    rows.append({
                        "Folder": fid,
                        "Version": filename_ref or "",
                        # "Typ": process_python_CC,
                        "Distances Path": py_dist_path,
                        "Params Path": py_params_path if os.path.exists(py_params_path) else "",
                        **stats
                    })

            # ----- CloudCompare
            if process_python_CC == "CC":
                cc_path        = cls._resolve(fid, f"CC_{filename_ref}_m3c2_distances.txt")
                cc_params_path = cls._resolve(fid, f"CC_{filename_ref}_m3c2_params.txt")
                if os.path.exists(cc_path):
                    try:
                        df = pd.read_csv(cc_path, sep=";")
                        col = "M3C2 distance"
                        if col in df.columns:
                            values = df[col].astype(float).values
                            stats = cls.calc_stats(
                                values,
                                params_path=cc_params_path if os.path.exists(cc_params_path) else None,
                                bins=bins,
                                range_override=range_override,
                                min_expected=min_expected,
                                outlier_multiplicator=outlier_multiplicator,
                                outlier_method=outlier_method
                            )
                            rows.append({
                                "Folder": fid,
                                "Version": filename_ref or "",
                                # "Typ": process_python_CC,
                                "Distances Path": cc_path,
                                "Params Path": cc_params_path if os.path.exists(cc_params_path) else "",
                                **stats
                            })
                        else:
                            print(f"[Stats] Spalte '{col}' fehlt in: {cc_path}")
                    except Exception as e:
                        print(f"[Stats] Konnte CC-Datei nicht lesen für {fid}: {e}")

        df_result = pd.DataFrame(rows)

        if out_path and not df_result.empty:
            if output_format.lower() == "json":
                cls._append_df_to_json(df_result, out_path)
            else:
                cls._append_df_to_excel(df_result, out_path, sheet_name=sheet_name)

        return df_result


    @staticmethod
    def write_table(
        rows: List[Dict],
        out_path: str = "m3c2_stats_all.xlsx",
        sheet_name: str = "Results",
        output_format: str = "excel",
    ) -> None:
        df = pd.DataFrame(rows)
        if df.empty:
            return
        if output_format.lower() == "json":
            StatisticsService._append_df_to_json(df, out_path)
        else:
            StatisticsService._append_df_to_excel(df, out_path, sheet_name=sheet_name)



    # =============================
    # Helpers
    # =============================

    @staticmethod
    def get_outlier_mask(clipped, method, outlier_multiplicator):
        if method == "rmse":
            rmse = np.sqrt(np.mean(clipped ** 2))
            outlier_threshold = (outlier_multiplicator * rmse)
            outlier_mask = np.abs(clipped) > outlier_threshold
        elif method == "iqr":
            q1 = np.percentile(clipped, 25)
            q3 = np.percentile(clipped, 75)
            iqr = q3 - q1
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            outlier_threshold = f"({lower_bound:.3f}, {upper_bound:.3f})"
            outlier_mask = (clipped < lower_bound) | (clipped > upper_bound)
        elif method == "std":
            mu = np.mean(clipped)
            std = np.std(clipped)
            outlier_threshold = (outlier_multiplicator * std)
            outlier_mask = np.abs(clipped - mu) > outlier_threshold
        elif method == "nmad":
            med  = np.median(clipped)
            nmad = 1.4826 * np.median(np.abs(clipped - med))
            outlier_threshold = (outlier_multiplicator * nmad)
            outlier_mask = np.abs(clipped - med) > outlier_threshold
        else:
            raise ValueError("Unbekannte Methode für Ausreißer-Erkennung: 'rmse', 'iqr', 'std', 'nmad'")
        return outlier_mask, outlier_threshold
    
    @staticmethod
    def _fit_distributions(
        clipped: np.ndarray,
        hist: np.ndarray,
        bin_edges: np.ndarray,
        min_expected: Optional[float],
    ) -> Dict[str, float]:
        """Fit Gaussian and Weibull distributions and compute Chi² metrics.

        Parameters
        ----------
        clipped:
            Werte, die innerhalb des gewählten Bereichs liegen.
        hist, bin_edges:
            Histogramm der ``clipped`` Werte und die zugehörigen Grenzen.
        min_expected:
            Mindestanzahl erwarteter Werte pro Bin für die Chi²-Berechnung.

        Returns
        -------
        dict
            Kennzahlen der Fits, inklusive Parameter und Chi²-Werten.
        """

        N = int(hist.sum())
        assert N == len(clipped), f"Histogram N ({N}) != len(clipped) ({len(clipped)})"

        # ------ Gauss ------ #
        mu, std = norm.fit(clipped)

        # Histogramm-Fit vorbereiten
        cdfL = norm.cdf(bin_edges[:-1], mu, std)  # CDF am linken Rand jedes Bins
        cdfR = norm.cdf(bin_edges[1:], mu, std)   # CDF am rechten Rand jedes Bins

        # Erwartete Häufigkeiten unter der Gauß-Verteilung
        expected_gauss = N * (cdfR - cdfL)

        # Kleine erwartete Werte aussortieren, um Division durch 0 zu vermeiden
        eps = 1e-12
        thr = min_expected if min_expected is not None else eps
        maskG = expected_gauss > thr

        # Pearson-Chi²
        pearson_gauss = float(
            np.sum((hist[maskG] - expected_gauss[maskG]) ** 2 / expected_gauss[maskG])
        )

        # ------ Weibull ------ #

        # Fit der Weibull-Verteilung
        a, loc, b = weibull_min.fit(clipped) # Shape, Scale, Location

        # Erwartete Häufigkeiten unter der Weibull-Verteilung
        cdfL = weibull_min.cdf(bin_edges[:-1], a, loc=loc, scale=b)
        cdfR = weibull_min.cdf(bin_edges[1:], a, loc=loc, scale=b)

        expected_weib = N * (cdfR - cdfL)

        # Kleine erwartete Klassen ausschließen
        maskW = expected_weib > thr

        # Pearson-Chi² für Weibull
        pearson_weib = float(
            np.sum((hist[maskW] - expected_weib[maskW]) ** 2 / expected_weib[maskW])
        )

        skew_weibull = float(weibull_min(a, loc=loc, scale=b).stats(moments="s"))
        # Modus (nur definiert, wenn a > 1)
        mode_weibull = float(loc + b * ((a - 1) / a) ** (1 / a)) if a > 1 else float(loc)

        return {
            "mu": float(mu),
            "std": float(std),
            "pearson_gauss": pearson_gauss,
            "a": float(a),
            "loc": float(loc),
            "b": float(b),
            "pearson_weib": pearson_weib,
            "skew_weibull": skew_weibull,
            "mode_weibull": mode_weibull,
        }

    @staticmethod
    def _compute_outliers(
        inliers: np.ndarray, outliers: np.ndarray
    ) -> Dict[str, float]:
        """Bestimme Kennzahlen zu Inliern und Outliern.

        Erwartet bereits separierte Arrays ``inliers`` und ``outliers``.
        """

        mean_out = float(np.mean(outliers)) if outliers.size else np.nan
        std_out = float(np.std(outliers)) if outliers.size > 0 else np.nan

        pos_out = int(np.sum(outliers > 0))
        neg_out = int(np.sum(outliers < 0))
        pos_in = int(np.sum(inliers > 0))
        neg_in = int(np.sum(inliers < 0))

        return {
            "outlier_count": int(outliers.size),
            "inlier_count": int(inliers.size),
            "mean_out": mean_out,
            "std_out": std_out,
            "pos_out": pos_out,
            "neg_out": neg_out,
            "pos_in": pos_in,
            "neg_in": neg_in,
        }

    @staticmethod
    def _basic_stats(values: np.ndarray, tolerance: float) -> Dict[str, float]:
        """Berechne Grundkennzahlen für ein Werte-Array.

        Gibt ein Dictionary mit Summen, Momenten und weiteren Kennzahlen
        zurück. Bei leeren Arrays werden ``np.nan`` bzw. 0 geliefert.
        """

        if values.size == 0:
            return {
                "Valid Count": 0,
                "Valid Sum": 0.0,
                "Valid Squared Sum": 0.0,
                "Min": np.nan,
                "Max": np.nan,
                "Mean": np.nan,
                "Median": np.nan,
                "RMS": np.nan,
                "Std Empirical": np.nan,
                "MAE": np.nan,
                "NMAD": np.nan,
                "Q05": np.nan,
                "Q25": np.nan,
                "Q75": np.nan,
                "Q95": np.nan,
                "IQR": np.nan,
                "Skewness": np.nan,
                "Kurtosis": np.nan,
                "Anteil |Distanz| > 0.01": np.nan,
                "Anteil [-2Std,2Std]": np.nan,
                "Max |Distanz|": np.nan,
                "Bias": np.nan,
                "Within-Tolerance": np.nan,
                "Jaccard Index": np.nan,
                "Dice Coefficient": np.nan,
            }

        valid_count = int(values.size)
        valid_sum = float(np.sum(values))
        valid_squared_sum = float(np.sum(values ** 2))
        min_val = float(np.min(values))
        max_val = float(np.max(values))
        mean_val = float(np.mean(values))
        median_val = float(np.median(values))
        rms_val = float(np.sqrt(np.mean(values ** 2)))
        std_emp = float(np.std(values))
        mae = float(np.mean(np.abs(values)))
        mad = float(np.median(np.abs(values - median_val)))
        nmad = float(1.4826 * mad)
        q05 = float(np.percentile(values, 5))
        q25 = float(np.percentile(values, 25))
        q75 = float(np.percentile(values, 75))
        q95 = float(np.percentile(values, 95))
        iqr = float(q75 - q25)
        skew = float(pd.Series(values).skew())
        kurt = float(pd.Series(values).kurt())
        share_abs_gt = float(np.mean(np.abs(values) > 0.01))
        share_2std = float(np.mean((values > -2 * std_emp) & (values < 2 * std_emp)))
        max_abs = float(np.max(np.abs(values)))
        bias = mean_val
        within_tolerance = float(np.mean(np.abs(values) <= tolerance))
        intersection = np.sum((values > -tolerance) & (values < tolerance))
        union = len(values)
        jaccard_index = intersection / union if union > 0 else np.nan
        dice_coefficient = (2 * intersection) / (2 * union) if union > 0 else np.nan

        return {
            "Valid Count": valid_count,
            "Valid Sum": valid_sum,
            "Valid Squared Sum": valid_squared_sum,
            "Min": min_val,
            "Max": max_val,
            "Mean": mean_val,
            "Median": median_val,
            "RMS": rms_val,
            "Std Empirical": std_emp,
            "MAE": mae,
            "NMAD": nmad,
            "Q05": q05,
            "Q25": q25,
            "Q75": q75,
            "Q95": q95,
            "IQR": iqr,
            "Skewness": skew,
            "Kurtosis": kurt,
            "Anteil |Distanz| > 0.01": share_abs_gt,
            "Anteil [-2Std,2Std]": share_2std,
            "Max |Distanz|": max_abs,
            "Bias": bias,
            "Within-Tolerance": within_tolerance,
            "Jaccard Index": jaccard_index,
            "Dice Coefficient": dice_coefficient,
        }

    @staticmethod
    def _load_params(params_path: Optional[str]) -> Tuple[float, float]:
        """Lese ``NormalScale`` und ``SearchScale`` aus einer Parameterdatei.

        Returns ``(np.nan, np.nan)`` falls die Datei nicht existiert oder die
        Werte fehlen.
        """

        normal_scale = np.nan
        search_scale = np.nan
        if params_path and os.path.exists(params_path):
            with open(params_path, "r") as f:
                for line in f:
                    if line.startswith("NormalScale="):
                        normal_scale = float(line.strip().split("=")[1])
                    elif line.startswith("SearchScale="):
                        search_scale = float(line.strip().split("=")[1])
        return normal_scale, search_scale

    @staticmethod
    def _resolve(fid: str, filename: str) -> str:
        p1 = os.path.join(fid, filename)
        if os.path.exists(p1):
            return p1
        return os.path.join("data", fid, filename)

    @staticmethod
    def _now_timestamp() -> str:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


    @staticmethod
    def _append_df_to_excel(df_new: pd.DataFrame, out_xlsx: str, sheet_name: str = "Results") -> None:
        """
        Hängt ``df_new`` an eine bestehende Excel-Datei an (oder erzeugt sie),
        sorgt für ``Timestamp`` als erste Spalte und harmonisiert Spalten.
        """

        if df_new is None or df_new.empty:
            return

        # Timestamp-Spalte einfügen (erste Spalte)
        ts = StatisticsService._now_timestamp()
        df_new = df_new.copy()
        df_new.insert(0, "Timestamp", ts)

        try:
            from openpyxl import load_workbook, Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ModuleNotFoundError as e:
            raise ModuleNotFoundError(
                "Zum Schreiben nach Excel wird 'openpyxl' benötigt. Bitte installieren: pip install openpyxl"
            ) from e

        # Daten in kanonische Reihenfolge bringen, fehlende Spalten mit NaN auffüllen,
        # aber zusätzliche Spalten behalten (kommen am Ende).
        original_cols = list(df_new.columns)
        for c in CANONICAL_COLUMNS:
            if c not in df_new.columns:
                df_new[c] = np.nan
        extra_cols = [c for c in original_cols if c not in CANONICAL_COLUMNS]
        df_new = df_new[CANONICAL_COLUMNS + extra_cols]

        # Ensure output directory exists
        out_dir = os.path.dirname(out_xlsx)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)


        if os.path.exists(out_xlsx):
            wb = load_workbook(out_xlsx)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

            # In manchen Workbooks enthält die erste Zeile gruppierte Überschriften
            # (teilweise mit zusammengeführten Zellen). Die echten Spaltennamen
            # stehen in Zeile 2.
            header_row = 2 if ws.max_row >= 2 else 1
            existing_cols = [cell.value for cell in ws[header_row] if cell.value is not None]
            if not existing_cols:
                # Kein Header vorhanden -> Standard-Header in Zeile 2 schreiben
                for idx, col in enumerate(CANONICAL_COLUMNS, start=1):
                    ws.cell(row=header_row, column=idx, value=col)
                existing_cols = CANONICAL_COLUMNS.copy()

            # Neue Spalten, die noch nicht im Sheet existieren, anhängen
            for c in df_new.columns:
                if c not in existing_cols:
                    existing_cols.append(c)
                    ws.cell(row=header_row, column=len(existing_cols), value=c)

            df_new = df_new.reindex(columns=existing_cols)
            for row in dataframe_to_rows(df_new, index=False, header=False):
                ws.append(row)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(df_new.columns.tolist())
            for row in dataframe_to_rows(df_new, index=False, header=False):
                ws.append(row)

        wb.save(out_xlsx)

    @staticmethod
    def _append_df_to_json(df_new: pd.DataFrame, out_json: str) -> None:
        """Wie ``_append_df_to_excel``, aber schreibt eine JSON-Datei."""
        if df_new is None or df_new.empty:
            return

        ts = StatisticsService._now_timestamp()
        df_new = df_new.copy()
        df_new.insert(0, "Timestamp", ts)

        # Ensure output directory exists
        out_dir = os.path.dirname(out_json)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)

        if os.path.exists(out_json):
            try:
                df_old = pd.read_json(out_json)
            except Exception:
                df_old = pd.DataFrame(columns=["Timestamp"])

            cols = list(df_old.columns) if not df_old.empty else ["Timestamp"]
            if "Timestamp" not in cols:
                cols.insert(0, "Timestamp")
            for c in df_new.columns:
                if c not in cols:
                    cols.append(c)

            df_old = df_old.reindex(columns=cols)
            df_new = df_new.reindex(columns=cols)
            df_all = pd.concat([df_old, df_new], ignore_index=True)
        else:
            df_all = df_new

        for c in CANONICAL_COLUMNS:
            if c not in df_all.columns:
                df_all[c] = np.nan
        df_all = df_all.reindex(columns=CANONICAL_COLUMNS)

        df_all.to_json(out_json, orient="records", indent=2)


    # --------------------------------------------- #
    # SINGLE-CLOUD STATISTIKEN (ohne Distanzwerte)
    # --------------------------------------------- #

    @staticmethod
    def _calc_single_cloud_stats(
        points: np.ndarray,
        area_m2: Optional[float] = None,
        radius: float = 1.0,
        k: int = 6,
        sample_size: Optional[int] = 100_000,
        use_convex_hull: bool = True,
    ) -> Dict:
        """Berechne Qualitätsmetriken für eine Punktwolke.

        Diese interne Funktion erwartet bereits geladene Punkte und liefert die
        statistischen Kennzahlen zurück. Die öffentliche Variante
        :meth:`calc_single_cloud_stats` kümmert sich zusätzlich um das Laden der
        Daten sowie das Schreiben in die Ergebnisdatei.
        """
        if points is None or len(points) == 0:
            raise ValueError("Points array is empty")
        P = np.asarray(points, dtype=float)
        if P.ndim != 2 or P.shape[1] != 3:
            raise ValueError("points must be of shape (N, 3)")

        # Globale Z-Statistik
        z = P[:, 2]
        num = len(P)
        z_min, z_max = float(np.min(z)), float(np.max(z))
        z_mean = float(np.mean(z))
        z_median = float(np.median(z))
        z_std = float(np.std(z))
        z_q05, z_q25, z_q75, z_q95 = map(float, np.percentile(z, [5, 25, 75, 95]))

        # XY-Fläche
        xy = P[:, :2]
        if area_m2 is None:
            area_bbox = StatisticsService._bbox_area_xy(xy)
            area_hull = StatisticsService._convex_hull_area_xy(xy) if use_convex_hull else np.nan
            area_m2_est = area_hull if use_convex_hull and not np.isnan(area_hull) else area_bbox
            area_used = float(area_m2_est)
            area_src = "convex_hull" if use_convex_hull and not np.isnan(area_hull) else "bbox"
        else:
            area_used = float(area_m2)
            area_src = "given"

        # Globale Dichte
        density_global = float(num / area_used) if area_used > 0 else np.nan

        # Subsample für lokale Metriken
        idx = np.arange(num)
        if sample_size and num > sample_size:
            idx = np.random.choice(num, size=sample_size, replace=False)
        S = P[idx]

        # kNN-Abstände
        nn = NearestNeighbors(n_neighbors=min(k + 1, len(S))).fit(S)
        dists_knn, _ = nn.kneighbors(S)
        mean_nn_all = float(np.mean(dists_knn[:, 1:]))
        mean_nn_kth = float(np.mean(dists_knn[:, min(k, dists_knn.shape[1]-1)]))

        # Radius-Nachbarschaften
        nbrs = NearestNeighbors(radius=radius).fit(S)
        ind_list = nbrs.radius_neighbors(S, return_distance=False)
        vol = 4.0 / 3.0 * np.pi * (radius ** 3)
        local_dens: List[float] = []
        rough: List[float] = []
        lin_list: List[float] = []
        pla_list: List[float] = []
        sph_list: List[float] = []
        anis_list: List[float] = []
        omni_list: List[float] = []
        eigent_list: List[float] = []
        curv_list: List[float] = []
        vert_list: List[float] = []
        normals: List[np.ndarray] = []

        for ind in ind_list:
            if ind.size < 3:
                continue
            neigh = S[ind]
            local_dens.append(ind.size / vol)
            c = np.mean(neigh, axis=0)
            U = neigh - c
            pca = PCA(n_components=3).fit(U)
            n = pca.components_[-1]
            d = np.abs(U @ n)
            rough.append(float(np.std(d)))
            evals = np.sort(pca.explained_variance_)[::-1]
            if evals[0] <= 0:
                continue
            linearity = (evals[0] - evals[1]) / evals[0]
            planarity = (evals[1] - evals[2]) / evals[0]
            sphericity = evals[2] / evals[0]
            anisotropy = (evals[0] - evals[2]) / evals[0]
            omnivariance = float(np.cbrt(np.prod(evals)))
            sum_eval = float(np.sum(evals))
            if sum_eval > 0:
                ratios = evals / sum_eval
                eigenentropy = float(-np.sum(ratios * np.log(ratios + 1e-15)))
                curvature = float(evals[2] / sum_eval)
            else:
                eigenentropy = np.nan
                curvature = np.nan
            verticality = float(
                np.degrees(np.arccos(np.clip(np.abs(n[2]), -1.0, 1.0)))
            )
            lin_list.append(float(linearity))
            pla_list.append(float(planarity))
            sph_list.append(float(sphericity))
            anis_list.append(float(anisotropy))
            omni_list.append(omnivariance)
            eigent_list.append(eigenentropy)
            curv_list.append(curvature)
            vert_list.append(verticality)
            normals.append(n)

        def _agg(arr):
            if len(arr) == 0:
                return (np.nan, np.nan, np.nan, np.nan)
            a = np.asarray(arr, dtype=float)
            return (
                float(np.mean(a)),
                float(np.median(a)),
                float(np.percentile(a, 5)),
                float(np.percentile(a, 95)),
            )

        dens_mean, dens_med, dens_q05, dens_q95 = _agg(local_dens)
        rough_mean, rough_med, rough_q05, rough_q95 = _agg(rough)
        lin_mean, lin_med, _, _ = _agg(lin_list)
        pla_mean, pla_med, _, _ = _agg(pla_list)
        sph_mean, sph_med, _, _ = _agg(sph_list)
        anis_mean, anis_med, _, _ = _agg(anis_list)
        omni_mean, omni_med, _, _ = _agg(omni_list)
        eig_mean, eig_med, _, _ = _agg(eigent_list)
        curv_mean, curv_med, _, _ = _agg(curv_list)
        vert_mean, vert_med, vert_q05, vert_q95 = _agg(vert_list)

        # Normalenkonsistenz
        normal_std_deg = np.nan
        if len(normals) > 3:
            N = np.asarray(normals)
            mean_n = np.mean(N, axis=0)
            if np.linalg.norm(mean_n) > 0:
                mean_n = mean_n / np.linalg.norm(mean_n)
            for i in range(N.shape[0]):
                if np.dot(N[i], mean_n) < 0:
                    N[i] = -N[i]
            cosang = np.clip(N @ mean_n, -1.0, 1.0)
            ang = np.degrees(np.arccos(cosang))
            normal_std_deg = float(np.std(ang))

        return {
            "Num Points": num,
            "Area Source": area_src,
            "Area XY [m^2]": area_used,
            "Density Global [pt/m^2]": density_global,
            "Z Min": z_min,
            "Z Max": z_max,
            "Z Mean": z_mean,
            "Z Median": z_median,
            "Z Std": z_std,
            "Z Q05": z_q05,
            "Z Q25": z_q25,
            "Z Q75": z_q75,
            "Z Q95": z_q95,
            f"Mean NN Dist (1..{k})": mean_nn_all,
            f"Mean Dist to {k}-NN": mean_nn_kth,
            "Local Density Mean [pt/m^3]": dens_mean,
            "Local Density Median [pt/m^3]": dens_med,
            "Local Density Q05 [pt/m^3]": dens_q05,
            "Local Density Q95 [pt/m^3]": dens_q95,
            "Roughness Mean [m]": rough_mean,
            "Roughness Median [m]": rough_med,
            "Roughness Q05 [m]": rough_q05,
            "Roughness Q95 [m]": rough_q95,
            "Linearity Mean": lin_mean,
            "Linearity Median": lin_med,
            "Planarity Mean": pla_mean,
            "Planarity Median": pla_med,
            "Sphericity Mean": sph_mean,
            "Sphericity Median": sph_med,
            "Anisotropy Mean": anis_mean,
            "Anisotropy Median": anis_med,
            "Omnivariance Mean": omni_mean,
            "Omnivariance Median": omni_med,
            "Eigenentropy Mean": eig_mean,
            "Eigenentropy Median": eig_med,
            "Curvature Mean": curv_mean,
            "Curvature Median": curv_med,
            "Verticality Mean [deg]": vert_mean,
            "Verticality Median [deg]": vert_med,
            "Verticality Q05 [deg]": vert_q05,
            "Verticality Q95 [deg]": vert_q95,
            "Normal Std Angle [deg]": normal_std_deg,
            "Radius [m]": float(radius),
            "k-NN": int(k),
            "Sampled Points": int(len(S)),
        }

    @staticmethod
    def _bbox_area_xy(xy: np.ndarray) -> float:
        x_min, y_min = np.min(xy[:, 0]), np.min(xy[:, 1])
        x_max, y_max = np.max(xy[:, 0]), np.max(xy[:, 1])
        return float((x_max - x_min) * (y_max - y_min))

    @staticmethod
    def _convex_hull_area_xy(xy: np.ndarray) -> float:
        try:
            from scipy.spatial import ConvexHull
        except Exception:
            return np.nan
        hull = ConvexHull(xy)
        return float(hull.volume)

    # --------------------------------------------- #
    # High-Level API für Single-Cloud-Statistiken
    # --------------------------------------------- #

    @classmethod
    def calc_single_cloud_stats(
        cls,
        folder_ids: List[str],
        filename_mov: str = "mov",
        filename_ref: str = "ref",
        area_m2: Optional[float] = None,
        radius: float = 1.0,
        k: int = 6,
        sample_size: Optional[int] = 100_000,
        use_convex_hull: bool = True,
        out_path: str = "m3c2_stats_clouds.xlsx",
        sheet_name: str = "CloudStats",
        output_format: str = "excel",
    ) -> pd.DataFrame:
        """Berechne Single-Cloud-Kennzahlen und speichere sie.

        Parameters
        ----------
        folder_ids:
            Liste von Ordnern, in denen die Punktwolken liegen.
        filename_mov, filename_ref:
            Basisnamen der bewegten und Referenzwolke.
        area_m2, radius, k, sample_size, use_convex_hull:
            Parameter, die an :func:`_calc_single_cloud_stats` weitergereicht werden.
        out_path, sheet_name, output_format:
            Ziel-Datei und Format für die Ausgabe.
        """

        rows: List[Dict] = []

        for fid in folder_ids:
            ds = DataSource(fid, filename_mov, filename_ref)
            mov, ref, _ = ds.load_points()
            for fname, epoch in ((filename_mov, mov), (filename_ref, ref)):
                pts = epoch.cloud if hasattr(epoch, "cloud") else epoch
                stats = cls._calc_single_cloud_stats(
                    pts,
                    area_m2=area_m2,
                    radius=radius,
                    k=k,
                    sample_size=sample_size,
                    use_convex_hull=use_convex_hull,
                )
                stats.update({"File": fname, "Folder": fid})
                rows.append(stats)

        df_result = pd.DataFrame(rows)
        if out_path and rows:
            cls.write_cloud_stats(
                rows,
                out_path=out_path,
                sheet_name=sheet_name,
                output_format=output_format,
            )
        return df_result

    @staticmethod
    def write_cloud_stats(
        rows: List[Dict],
        out_path: str = "m3c2_stats_clouds.xlsx",
        sheet_name: str = "CloudStats",
        output_format: str = "excel",
    ) -> None:
        df = pd.DataFrame(rows)
        if df.empty:
            return
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        df.insert(0, "Timestamp", ts)
        if os.path.exists(out_path):
            try:
                if output_format.lower() == "json":
                    old = pd.read_json(out_path)
                else:
                    old = pd.read_excel(out_path, sheet_name=sheet_name)
            except Exception:
                old = pd.DataFrame(columns=["Timestamp"])
            cols = list(old.columns) if not old.empty else ["Timestamp"]
            for c in df.columns:
                if c not in cols:
                    cols.append(c)
            old = old.reindex(columns=cols)
            df = df.reindex(columns=cols)
            all_df = pd.concat([old, df], ignore_index=True)
        else:
            all_df = df

        if output_format.lower() == "json":
            all_df.to_json(out_path, orient="records", indent=2)
        else:
            with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as w:
                all_df.to_excel(w, index=False, sheet_name=sheet_name)
